# gestion/utils.py
# ══════════════════════════════════════════════════════════
#  SETRAM — Utilitaires backend
#  Pipeline Excel : basé sur extract_bulletins.py (validé)
# ══════════════════════════════════════════════════════════

import re
import datetime
from django.utils import timezone
from .models import Notification, HistoriqueModification, Bulletin, Course


# ── Notifications & historique ────────────────────────────────────────────────

def envoyer_notification(destinataire, type_notif, titre, message):
    """Crée une notification en base + push Firebase si fcm_token présent."""
    Notification.objects.create(
        destinataire=destinataire,
        type_notif=type_notif,
        titre=titre,
        message=message,
    )
    if getattr(destinataire, 'fcm_token', None):
        _envoyer_push_firebase(destinataire.fcm_token, titre, message)


def _envoyer_push_firebase(fcm_token, titre, message):
    try:
        import firebase_admin
        from firebase_admin import messaging, credentials
        if not firebase_admin._apps:
            cred = credentials.Certificate('chemin/vers/serviceAccountKey.json')
            firebase_admin.initialize_app(cred)
        msg = messaging.Message(
            notification=messaging.Notification(title=titre, body=message),
            token=fcm_token,
        )
        messaging.send(msg)
    except Exception as e:
        print(f"[FCM] Erreur : {e}")


def enregistrer_historique(utilisateur, action, table, objet_id, details=None):
    HistoriqueModification.objects.create(
        utilisateur=utilisateur,
        action=action,
        table_ciblee=table,
        objet_id=objet_id,
        details=details or {},
    )


# ── Configuration colonnes Excel ──────────────────────────────────────────────
# (identique à extract_bulletins.py — validé sur les vrais fichiers)
#   (group_idx, course_col, rame_col, origin_col, dep_col, dep_delay_col,
#    dest_col, arr_col, arr_delay_col)

_SERVICE_COLS = [
    (0,  0,  8,  3,  4,  5,  6,  None, None),   # services 01–10
    (1, 11, 19, 14, 15, 16, 17,   18,   19),     # services 11–20
    (2, 22, 30, 25, 26, 27, 28,   29,   30),     # services 21–30
    (3, 33, 41, 36, 37, 38, 39,   40,   41),     # services 31–40
    (4, 44, 52, 47, 48, 49, 50,   51,   52),     # services 41–50
    (5, 55, 63, 58, 59, 60, 61,   62,   63),     # services 51–60
    (6, 66, 74, 69, 70, 71, 72,   73,   74),     # services 61–70
]

_HEADER_KW = [
    'origine', 'heure', 'destination', 'retard', 'fin-service',
    'heures réelles', 'course', 'numéro', 'svp', 'prise', 'réelles',
]

_COURSE_SKIP = [
    'course', 'numéro', 'origine', 'heure', 'retard',
    'destination', 'service', 'fin-service', 'prise', 'svp',
]

_SHEET_TYPE_MAP = {
    'JO':       Bulletin.JO,
    'JS ET JV': Bulletin.JS_JV,
    'JS/JV':    Bulletin.JS_JV,
    'JS':       Bulletin.JS_JV,
    'JV':       Bulletin.JS_JV,
}


# ── Helpers internes ──────────────────────────────────────────────────────────

def _fmt_time(v):
    if isinstance(v, datetime.time):     return v
    if isinstance(v, datetime.datetime): return v.time()
    return None


def _clean_str(v):
    if v is None: return None
    s = str(v).strip()
    return None if s in ("#####", "") else s


def _is_header(v):
    if v is None: return False
    return any(kw in str(v).lower() for kw in _HEADER_KW)


def _safe_get(row, col):
    if col is None or col >= len(row): return None
    return row[col]


def _normaliser_station(val):
    if not val: return ''
    s = str(val).strip()
    mapping = {
        'depot': 'Dépôt', 'dépôt': 'Dépôt', 'dep': 'Dépôt',
        'bam': 'BAM', 'uam': 'UAM', 'khe': 'KHE',
        'zou1': 'ZOU1', 'zou2': 'ZOU2', 'palma': 'PALMA',
    }
    return mapping.get(s.lower(), s.upper())


def _detect_type_jour(sheet_name):
    """Détecte le type de jour depuis le nom de la feuille."""
    n = sheet_name.strip().upper()
    for key, val in _SHEET_TYPE_MAP.items():
        if n.startswith(key):
            return val
    return Bulletin.JO


# ── Extraction d'une feuille ──────────────────────────────────────────────────

def _extraire_feuille(ws):
    """
    Extrait tous les services d'une feuille Excel.
    Retourne une liste de dicts :
      { service_id, type_jour, heure_debut, heure_fin, courses: [...] }

    Logique identique à extract_bulletins.py, validée sur les vrais fichiers.
    """
    rows = list(ws.iter_rows(values_only=True))
    type_jour = _detect_type_jour(ws.title)

    # Trouver les débuts de blocs (colonne 0 contient "service NN")
    block_starts = [
        i for i, row in enumerate(rows)
        if row and isinstance(row[0], str)
        and re.match(r'service\s+\d+', row[0].lower())
    ]

    services = []

    for bi, bs in enumerate(block_starts):
        be = block_starts[bi + 1] if bi + 1 < len(block_starts) else bs + 20
        header_row = rows[bs]

        for (_, course_col, rame_col, origin_col,
             dep_col, dep_delay_col, dest_col,
             arr_col, arr_delay_col) in _SERVICE_COLS:

            # Identifier le service dans ce rang de colonnes
            svc_val = _safe_get(header_row, course_col)
            if not isinstance(svc_val, str) or not svc_val.lower().startswith('service'):
                continue

            m = re.search(r'(\d+)', svc_val)
            if not m:
                continue
            service_num = int(m.group(1))

            # Extraire les courses (offset +6 dans le bloc)
            courses_data = []
            current_course = None

            for ri in range(bs + 6, be):
                if ri >= len(rows):
                    break
                row = rows[ri]

                # Mettre à jour le nom de course courant
                cv = _clean_str(_safe_get(row, course_col))
                if cv and not any(kw in cv.lower() for kw in _COURSE_SKIP):
                    current_course = cv

                origin   = _clean_str(_safe_get(row, origin_col))
                dep_time = _fmt_time(_safe_get(row, dep_col))
                dest     = _clean_str(_safe_get(row, dest_col))
                arr_time = _fmt_time(_safe_get(row, arr_col))

                # Ignorer lignes vides
                if all(v is None for v in [origin, dep_time, dest]):
                    continue

                # Ignorer lignes d'en-tête résiduelles
                if any(_is_header(v) for v in [origin, dep_time, dest]):
                    continue

                # Ignorer heure de pause (00:30)
                if dep_time and dep_time == datetime.time(0, 30):
                    continue

                if not origin or not dep_time:
                    continue

                courses_data.append({
                    'numero_course':     current_course or '',
                    'origine':           _normaliser_station(origin),
                    'destination':       _normaliser_station(dest) if dest else '',
                    'heure_depart_prev': dep_time,
                    'heure_arrivee_prev': arr_time,
                })

            services.append({
                'service_num': service_num,
                'type_jour':   type_jour,
                'heure_debut': courses_data[0]['heure_depart_prev'] if courses_data else None,
                'heure_fin':   courses_data[-1]['heure_depart_prev'] if courses_data else None,
                'courses':     courses_data,
                'est_vide':    len(courses_data) == 0,
            })

    return services


# ── Point d'entrée principal ──────────────────────────────────────────────────

def importer_bulletins_excel(fichier, importeur):
    """
    Importe les bulletins depuis un fichier Excel Django UploadedFile.
    Utilisé par l'API (POST /bulletins/importer/) et l'admin Django.

    Retourne un dict rapport :
      { total, crees, mis_a_jour, vides, erreurs }
    """
    import openpyxl
    rapport = {
        'total': 0, 'crees': 0, 'mis_a_jour': 0,
        'vides': 0, 'erreurs': []
    }

    try:
        wb = openpyxl.load_workbook(fichier, data_only=True)
    except Exception as e:
        rapport['erreurs'].append(f"Impossible d'ouvrir le fichier : {e}")
        return rapport

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            services = _extraire_feuille(ws)
        except Exception as e:
            rapport['erreurs'].append(f"Erreur feuille '{sheet_name}': {e}")
            continue

        for svc in services:
            try:
                bulletin, created = Bulletin.objects.update_or_create(
                    numero=svc['service_num'],
                    type_jour=svc['type_jour'],
                    defaults={
                        'heure_debut':    svc['heure_debut'],
                        'heure_fin':      svc['heure_fin'],
                        'importe_par':    importeur,
                        'fichier_source': getattr(fichier, 'name', str(fichier)),
                    }
                )

                if created:
                    rapport['crees'] += 1
                else:
                    bulletin.courses.all().delete()
                    rapport['mis_a_jour'] += 1

                # Créer les courses en batch
                if svc['courses']:
                    Course.objects.bulk_create([
                        Course(
                            bulletin=bulletin,
                            numero_course=c['numero_course'],
                            origine=c['origine'],
                            destination=c['destination'],
                            heure_depart_prev=c['heure_depart_prev'],
                            heure_arrivee_prev=c['heure_arrivee_prev'],
                            ordre=i,
                        )
                        for i, c in enumerate(svc['courses'])
                    ])
                    rapport['total'] += 1
                else:
                    rapport['vides'] += 1

            except Exception as e:
                rapport['erreurs'].append(
                    f"Erreur service {svc['service_num']} ({sheet_name}): {e}")

    return rapport